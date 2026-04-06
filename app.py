from flask import Flask, request, jsonify, send_from_directory, session
import json
import os
import openpyxl
from openpyxl import load_workbook
from datetime import datetime

app = Flask(__name__, static_folder='static')
app.secret_key = 'your-secret-key-change-this-in-production'

RESULTS_FILE = 'results.xlsx'
ADMIN_PASSWORD = 'Behzod5664@'  # Change this!

QUESTIONS = [
    {"id": 1, "question": "Hozirgi O‘zbekiston hududida joylashgan qadimgi sivilizatsiya qaysi?", "options": ["Mesopotamiya", "So‘g‘diyona", "Misr", "Hind vodiysi"], "answer": "So‘g‘diyona"},
    {"id": 2, "question": "Samarqand qaysi savdo yo‘lining muhim markazi bo‘lgan?", "options": ["Ziravorlar yo‘li", "Ipak yo‘li", "Qahrabo yo‘li", "Ot-cho‘y yo‘li"], "answer": "Ipak yo‘li"},
    {"id": 3, "question": "Temuriylar davlatiga kim asos solgan?", "options": ["Chingizxon", "Bobur", "Amir Temur", "Ulug‘bek"], "answer": "Amir Temur"},
    {"id": 4, "question": "Amir Temur yana qanday nom bilan tanilgan?", "options": ["Aleksandr", "Tamerlan", "Sulaymon", "Attila"], "answer": "Tamerlan"},
    {"id": 5, "question": "Temurdan oldin hudud qaysi imperiya tarkibida bo‘lgan?", "options": ["Usmonli", "Mo‘g‘ullar", "Rim", "Fors"], "answer": "Mo‘g‘ullar"},
    {"id": 6, "question": "Ulug‘bek qaysi sohada mashhur bo‘lgan?", "options": ["Tibbiyot", "Astronomiya", "Me’morchilik", "Adabiyot"], "answer": "Astronomiya"},
    {"id": 7, "question": "Ulug‘bek rasadxonasi qayerda joylashgan?", "options": ["Buxoro", "Xiva", "Samarqand", "Toshkent"], "answer": "Samarqand"},
    {"id": 8, "question": "Bobur qayerda imperiya asos solgan?", "options": ["Xitoy", "Hindiston", "Eron", "Turkiya"], "answer": "Hindiston"},
    {"id": 9, "question": "Xiva xonligi markazi qaysi edi?", "options": ["Qo‘qon", "Buxoro", "Xiva", "Oltin O‘rda"], "answer": "Xiva"},
    {"id": 10, "question": "19-asrda O‘zbekiston qaysi imperiyaga qo‘shilgan?", "options": ["Britaniya", "Rossiya", "Usmonli", "Xitoy"], "answer": "Rossiya"},
    {"id": 11, "question": "Sovet davrida O‘zbekiston qanday nomlangan?", "options": ["Markaziy SSR", "O‘zbek SSR", "Turkiston SSR", "Sovet Osiyo"], "answer": "O‘zbek SSR"},
    {"id": 12, "question": "O‘zbekiston qachon mustaqillikka erishgan?", "options": ["1989", "1990", "1991", "1992"], "answer": "1991"},
    {"id": 13, "question": "Birinchi prezident kim bo‘lgan?", "options": ["Mirziyoyev", "Karimov", "Niyazov", "Nazarboyev"], "answer": "Karimov"},
    {"id": 14, "question": "O‘zbekiston poytaxti qaysi?", "options": ["Samarqand", "Buxoro", "Toshkent", "Xiva"], "answer": "Toshkent"},
    {"id": 15, "question": "Ichon Qal’a qaysi shaharda?", "options": ["Samarqand", "Xiva", "Buxoro", "Nukus"], "answer": "Xiva"},
    {"id": 16, "question": "So‘g‘diylar asosan nima bilan shug‘ullangan?", "options": ["Urush", "Savdo", "Dehqonchilik", "Chorvachilik"], "answer": "Savdo"},
    {"id": 17, "question": "Islomdan oldin keng tarqalgan din qaysi?", "options": ["Xristianlik", "Buddizm", "Zardushtiylik", "Hinduizm"], "answer": "Zardushtiylik"},
    {"id": 18, "question": "Islom dini qachon kirib kelgan?", "options": ["Yunonlar", "Arablar", "Mo‘g‘ullar", "Ruslar"], "answer": "Arablar"},
    {"id": 19, "question": "Ibn Sino qaysi sohada mashhur?", "options": ["Matematika", "Tibbiyot", "Astronomiya", "Adabiyot"], "answer": "Tibbiyot"},
    {"id": 20, "question": "Al-Xorazmiy nimaning asoschisi?", "options": ["Geometriya", "Algebra", "Fizika", "Kimyo"], "answer": "Algebra"},
    {"id": 21, "question": "Qaysi dengiz qurib bormoqda?", "options": ["Kaspiy", "Qora", "Orol", "O‘rta yer"], "answer": "Orol"},
    {"id": 22, "question": "Jadidlar harakati maqsadi nima?", "options": ["Urush", "Islohot", "Savdo", "Bosqin"], "answer": "Islohot"},
    {"id": 23, "question": "Sovet boshida o‘zbek yozuvi qaysi edi?", "options": ["Arab", "Kirill", "Lotin", "Xitoy"], "answer": "Lotin"},
    {"id": 24, "question": "Qaysi shahar ilm markazi bo‘lgan?", "options": ["Toshkent", "Buxoro", "Nukus", "Termiz"], "answer": "Buxoro"},
    {"id": 25, "question": "Oltin O‘rda qaysi imperiyaga tegishli?", "options": ["Usmonli", "Mo‘g‘ul", "Fors", "Rim"], "answer": "Mo‘g‘ul"},
    {"id": 26, "question": "Registon qayerda joylashgan?", "options": ["Buxoro", "Samarqand", "Xiva", "Andijon"], "answer": "Samarqand"},
    {"id": 27, "question": "Karimovdan keyin kim prezident bo‘lgan?", "options": ["Rahmon", "Mirziyoyev", "Putin", "Aliyev"], "answer": "Mirziyoyev"},
    {"id": 28, "question": "Movarounnahr nimani anglatadi?", "options": ["Kaspiy orti", "Ikki daryo oralig‘i", "Rossiya", "Xitoy"], "answer": "Ikki daryo oralig‘i"},
    {"id": 29, "question": "13-asr bosqini qaysi?", "options": ["Arab", "Mo‘g‘ul", "Fors", "Yunon"], "answer": "Mo‘g‘ul"},
    {"id": 30, "question": "Eng qadimiy shaharlardan biri qaysi?", "options": ["Nukus", "Samarqand", "Toshkent", "Andijon"], "answer": "Samarqand"},
]

def get_completed_users():
    if not os.path.exists(RESULTS_FILE):
        return set()
    wb = load_workbook(RESULTS_FILE)
    ws = wb.active
    completed = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            completed.add(row[0].lower().strip())
    return completed

def save_result(email, name, answers, score, total_time):
    if not os.path.exists(RESULTS_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Quiz Results"
        ws.append(["Email", "Full Name", "Score", "Total Questions", "Percentage", "Time Taken (s)", "Date", "Answers"])
    else:
        wb = load_workbook(RESULTS_FILE)
        ws = wb.active

    percentage = round((score / 30) * 100, 1)
    date_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    answers_str = json.dumps(answers)
    ws.append([email, name, score, 30, f"{percentage}%", total_time, date_str, answers_str])
    wb.save(RESULTS_FILE)

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/api/start', methods=['POST'])
def start_quiz():
    data = request.json
    email = data.get('email', '').lower().strip()
    name = data.get('name', '').strip()
    if not email or not name:
        return jsonify({'error': 'Email and name are required'}), 400
    completed = get_completed_users()
    if email in completed:
        return jsonify({'error': 'You have already completed this quiz. It can only be taken once.'}), 403
    session['email'] = email
    session['name'] = name
    return jsonify({'success': True, 'questions': QUESTIONS})

@app.route('/api/submit', methods=['POST'])
def submit_quiz():
    data = request.json
    email = session.get('email')
    name = session.get('name')
    if not email:
        return jsonify({'error': 'Session expired. Please restart.'}), 401
    completed = get_completed_users()
    if email in completed:
        return jsonify({'error': 'Already submitted.'}), 403
    answers = data.get('answers', {})
    total_time = data.get('totalTime', 0)
    score = 0
    result_details = {}
    for q in QUESTIONS:
        qid = str(q['id'])
        user_ans = answers.get(qid, '')
        correct = q['answer']
        is_correct = user_ans == correct
        if is_correct:
            score += 1
        result_details[qid] = {'userAnswer': user_ans, 'correct': correct, 'isCorrect': is_correct}
    save_result(email, name, result_details, score, total_time)
    session.clear()
    return jsonify({'success': True, 'score': score, 'total': 30, 'details': result_details})

@app.route('/api/admin/results')
def admin_results():
    password = request.args.get('password', '')
    if password != ADMIN_PASSWORD:
        return jsonify({'error': 'Unauthorized'}), 401
    if not os.path.exists(RESULTS_FILE):
        return jsonify({'results': []})
    wb = load_workbook(RESULTS_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        results.append(dict(zip(headers, row)))
    return jsonify({'results': results})

@app.route('/api/admin/download')
def download_excel():
    password = request.args.get('password', '')
    if password != ADMIN_PASSWORD:
        return jsonify({'error': 'Unauthorized'}), 401
    if not os.path.exists(RESULTS_FILE):
        return jsonify({'error': 'No results yet'}), 404
    return send_from_directory('.', RESULTS_FILE, as_attachment=True)

if __name__ == '__main__':
    os.makedirs('static', exist_ok=True)
    app.run(debug=True)
